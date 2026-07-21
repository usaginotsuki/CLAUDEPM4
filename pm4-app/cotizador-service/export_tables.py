#!/usr/bin/env python3
"""
export_tables.py — se ejecuta UNA VEZ en build time.
Lee el Excel con openpyxl (data_only=True) y exporta todas las tablas
de lookup a tables.json para que app.py las use sin necesitar scipy.
"""
import json, os, sys
import openpyxl

HERE     = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.normpath(os.path.join(HERE, '..', 'frontend', 'src', 'resources', 'cotizador2.xlsx'))
OUTPUT   = os.path.join(HERE, 'tables.json')

print(f'Leyendo {TEMPLATE}…')
# Abrimos el libro de Excel con los valores ya calculados
objWb = openpyxl.load_workbook(TEMPLATE, data_only=True)

dicTables = {}

# ── D&O ───────────────────────────────────────────────────────────────────────
objWs = objWb["Configuraciond&o"]

# Límites disponibles (fila 34, cols D-P)
lstDyoLimits = []
for intCol in range(4, 17):
    genValue = objWs.cell(34, intCol).value
    if genValue is not None:
        lstDyoLimits.append(genValue)

# Matriz de primas: fila 35-43, col C = facturación label, cols D-P = primas
lstDyoMatrix = []
for intRow in range(35, 44):
    strBilling = objWs.cell(intRow, 3).value
    if strBilling is None:
        continue
    lstPremiums = []
    for intCol in range(4, 4 + len(lstDyoLimits)):
        genValue = objWs.cell(intRow, intCol).value
        lstPremiums.append(None if (genValue is None or str(genValue) == 'N/A') else float(genValue))
    lstDyoMatrix.append({'fac': strBilling, 'primas': lstPremiums})

# NC/Anexo: fila 17-29, col B=límite, col C=límite_nc, col D=deducible
dicDyoNc = {}
for intRow in range(17, 30):
    intLimit = objWs.cell(intRow, 2).value
    intLimitNc = objWs.cell(intRow, 3).value
    intDeductible = objWs.cell(intRow, 4).value
    if intLimit is not None:
        dicDyoNc[str(int(intLimit))] = {
            'limite_nc': int(intLimitNc) if intLimitNc else None,
            'deducible': int(intDeductible)    if intDeductible    else None,
        }

dicTables['dyo'] = {
    'limites': [int(genLimit) for genLimit in lstDyoLimits],
    'matrix': lstDyoMatrix,
    'nc': dicDyoNc,
}
print(f'  D&O: {len(lstDyoMatrix)} filas × {len(lstDyoLimits)} límites')

# ── CC / Infidelidad ──────────────────────────────────────────────────────────
objWsCc = objWb['Nuevas Primas CYBER']

# Recorremos las filas para leer las primas de CC
dicCcPrimas = {}
for intRow in range(42, 600):
    genKey = objWsCc.cell(intRow, 5).value
    genValue = objWsCc.cell(intRow, 6).value
    if genKey is None:
        break
    if genValue is not None and str(genValue) not in ('N/A', '-', ''):
        dicCcPrimas[str(genKey)] = float(genValue)

# Leemos los deducibles de CC
dicCcDed = {}
for intRow in range(4, 12):
    genKey = objWsCc.cell(intRow, 22).value
    genValue = objWsCc.cell(intRow, 23).value
    if genKey is None:
        break
    dicCcDed[str(genKey)] = str(genValue) if genValue is not None else None

dicTables['cc'] = {'primas': dicCcPrimas, 'deducibles': dicCcDed}
print(f'  CC: {len(dicCcPrimas)} primas, {len(dicCcDed)} deducibles')

# ── PDySI / Cyber (v2: ENTRADAS B27-B30, SALIDAS R23-R25) ────────────────────
objWsCy = objWb['PRIMAS']

# Recorremos las filas para leer las primas de PDySI
dicCyPrimas = {}
for intRow in range(18, 60):
    genKey = objWsCy.cell(intRow, 1).value
    genValue = objWsCy.cell(intRow, 4).value
    if genKey is None:
        break
    if genValue is not None and str(genValue) not in ('N/A', '-', ''):
        try: dicCyPrimas[str(genKey)] = float(genValue)
        except: pass

# Leemos los deducibles de PDySI
dicCyDed = {}
for intRow in range(57, 100):
    genKey = objWsCy.cell(intRow, 1).value
    genValue = objWsCy.cell(intRow, 4).value
    if genKey is None:
        break
    if genValue is not None and str(genValue) not in ('N/A', '-', ''):
        try: dicCyDed[str(genKey)] = float(genValue)
        except: pass

dicTables['pdysi'] = {'primas': dicCyPrimas, 'deducibles': dicCyDed}
print(f'  PDySI: {len(dicCyPrimas)} primas, {len(dicCyDed)} deducibles')

# ── PI (v2: 3 alternativas, key = {fac}{deducible}{limite}) ───────────────────
dicPiTables = {}
for strSectorName, strSheetName in [('ABOGADOS', 'ABOGADOS'), ('ADMINISTRADORES', 'ADMIN PH'), ('CONTADORES', 'CONTADORES')]:
    objWsPi = objWb[strSheetName]
    # Construimos el lookup por llave para este sector
    dicLookup = {}
    for intRow in range(1, 800):
        genKey   = objWsPi.cell(intRow, 6).value   # col F = llave {fac}{ded}{lim}
        dblPrima = objWsPi.cell(intRow, 5).value   # col E = prima
        if genKey and dblPrima and str(dblPrima) not in ('N/A', '-', 'PRIMA'):
            try: dicLookup[str(genKey)] = float(dblPrima)
            except: pass
    dicPiTables[strSectorName] = dicLookup
    print(f'  PI/{strSectorName}: {len(dicLookup)} entradas')

dicTables['pi'] = dicPiTables

# Mapeo limit→deducible por sector (para auto-derivar deducible cuando no se pasa)
# Formato: {sector: {limite_str: deducible}}
dicPiDedMap = {}
for strSectorName, strSheetName in [('ABOGADOS', 'ABOGADOS'), ('ADMINISTRADORES', 'ADMIN PH'), ('CONTADORES', 'CONTADORES')]:
    objWsPi = objWb[strSheetName]
    # Guardamos el primer deducible encontrado por cada limite
    dicMapping = {}
    for intRow in range(1, 800):
        genKey   = objWsPi.cell(intRow, 6).value   # llave
        dblPrima = objWsPi.cell(intRow, 5).value
        intDeductible   = objWsPi.cell(intRow, 3).value   # col C = deducible
        intLimit   = objWsPi.cell(intRow, 4).value   # col D = limite
        if genKey and dblPrima and intDeductible and intLimit:
            try:
                strLimit = str(int(intLimit))
                if strLimit not in dicMapping:   # primer deducible encontrado para ese límite
                    dicMapping[strLimit] = int(intDeductible)
            except: pass
    dicPiDedMap[strSectorName] = dicMapping
    print(f'  PI deducibles/{strSectorName}: {len(dicMapping)} entradas')

dicTables['pi_ded_map'] = dicPiDedMap

# ── Guardar ───────────────────────────────────────────────────────────────────
# Guardamos todas las tablas en el archivo JSON final
with open(OUTPUT, 'w', encoding='utf-8') as objFile:
    json.dump(dicTables, objFile, ensure_ascii=False)

print(f'\nOK Exportado a {OUTPUT} ({os.path.getsize(OUTPUT) // 1024} KB)')
