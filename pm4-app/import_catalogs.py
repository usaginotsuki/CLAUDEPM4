#!/usr/bin/env python3
import json
import os
import sys
import uuid
import argparse
import openpyxl
import requests
import subprocess
import re

# Reconfigure stdout to use UTF-8 to prevent console encoding crashes on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Global mappings for catalog names to mock records
MOCK_RECORDS = {
    "cat-canal": [
        {"codigo": "5", "descripcion": "Centro atención telefónica"},
        {"codigo": "3", "descripcion": "Web"},
        {"codigo": "1", "descripcion": "Presencial"},
        {"codigo": "2", "descripcion": "Virtual"},
        {"codigo": "4", "descripcion": "Escrito"}
    ],
    "cat-tipo-id": [
        {"codigo": "CC", "descripcion": "Cédula de Ciudadanía"},
        {"codigo": "CE", "descripcion": "Cédula de Extranjería"},
        {"codigo": "NIT", "descripcion": "NIT"},
        {"codigo": "Pasaporte", "descripcion": "Pasaporte"},
        {"codigo": "TI", "descripcion": "Tarjeta de Identidad"}
    ],
    "cat-tipo-persona": [
        {"codigo": "1", "descripcion": "Natural"},
        {"codigo": "2", "descripcion": "Jurídica"}
    ],
    "cat-pais": [
        {"codigo": "170", "descripcion": "Colombia"},
        {"codigo": "840", "descripcion": "Estados Unidos"},
        {"codigo": "724", "descripcion": "España"},
        {"codigo": "032", "descripcion": "Argentina"},
        {"codigo": "484", "descripcion": "México"}
    ],
    "cat-dpto": [
        {"codigo_departamento": "ANTIOQUIA", "nombre_departamento": "Antioquia"},
        {"codigo_departamento": "ATLANTICO", "nombre_departamento": "Atlántico"},
        {"codigo_departamento": "BOGOTA", "nombre_departamento": "Bogotá D.C."},
        {"codigo_departamento": "BOLIVAR", "nombre_departamento": "Bolívar"},
        {"codigo_departamento": "CUNDINAMARCA", "nombre_departamento": "Cundinamarca"},
        {"codigo_departamento": "MAGDALENA", "nombre_departamento": "Magdalena"},
        {"codigo_departamento": "RISARALDA", "nombre_departamento": "Risaralda"},
        {"codigo_departamento": "SANTANDER", "nombre_departamento": "Santander"},
        {"codigo_departamento": "VALLE", "nombre_departamento": "Valle del Cauca"}
    ],
    "cat-mpio": [
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "MEDELLIN", "nombre_municipio": "Medellín"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "BELLO", "nombre_municipio": "Bello"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "ENVIGADO", "nombre_municipio": "Envigado"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "ITAGUI", "nombre_municipio": "Itagüí"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "RIONEGRO", "nombre_municipio": "Rionegro"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "BARRANQUILLA", "nombre_municipio": "Barranquilla"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "SOLEDAD", "nombre_municipio": "Soledad"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "MALAMBO", "nombre_municipio": "Malambo"},
        {"codigo_departamento": "BOGOTA", "codigo_municipio": "BOGOTA", "nombre_municipio": "Bogotá D.C."},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "CARTAGENA", "nombre_municipio": "Cartagena"},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "TURBACO", "nombre_municipio": "Turbaco"},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "MAGANGUE", "nombre_municipio": "Magangué"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "SOACHA", "nombre_municipio": "Soacha"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "CHIA", "nombre_municipio": "Chía"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "CAJICA", "nombre_municipio": "Cajicá"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "FACATATIVA", "nombre_municipio": "Facatativá"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "SANTA_MARTA", "nombre_municipio": "Santa Marta"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "CIENAGA", "nombre_municipio": "Ciénaga"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "FUNDACION", "nombre_municipio": "Fundación"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "PEREIRA", "nombre_municipio": "Pereira"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "DOSQUEBRADAS", "nombre_municipio": "Dosquebradas"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "SANTA_ROSA", "nombre_municipio": "Santa Rosa de Cabal"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "BUCARAMANGA", "nombre_municipio": "Bucaramanga"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "FLORIDABLANCA", "nombre_municipio": "Floridablanca"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "GIRON", "nombre_municipio": "Girón"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "CALI", "nombre_municipio": "Cali"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "PALMIRA", "nombre_municipio": "Palmira"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "BUENAVENTURA", "nombre_municipio": "Buenaventura"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "TULUA", "nombre_municipio": "Tuluá"}
    ],
    "cat-producto-sfc": [
        {"codigo_producto_sfc": "101", "nombre_producto_sfc": "Automóviles"},
        {"codigo_producto_sfc": "102", "nombre_producto_sfc": "Vida"},
        {"codigo_producto_sfc": "103", "nombre_producto_sfc": "Hogar"},
        {"codigo_producto_sfc": "104", "nombre_producto_sfc": "Salud"},
        {"codigo_producto_sfc": "105", "nombre_producto_sfc": "Cumplimiento"}
    ],
    "cat-motivo-sfc": [
        {"codigo": "301", "descripcion": "No pago siniestro"},
        {"codigo": "302", "descripcion": "Demora pago"},
        {"codigo": "303", "descripcion": "Mala atención"},
        {"codigo": "304", "descripcion": "Cobro indebido"},
        {"codigo": "305", "descripcion": "Incumplimiento de contrato"}
    ],
    "cat-tipo-sol": [
        {"codigo": "QUEJA_DIRECTA", "descripcion": "Queja Directa SmartSupervision"},
        {"codigo": "REQUERIMIENTO", "descripcion": "Requerimiento"},
        {"codigo": "SUGERENCIA", "descripcion": "Sugerencia"},
        {"codigo": "FELICITACION", "descripcion": "Felicitación"}
    ],
    "cat-instancia": [
        {"codigo": "2", "descripcion": "Entidad vigilada"},
        {"codigo": "1", "descripcion": "Defensor CF"},
        {"codigo": "3", "descripcion": "SFC"}
    ],
    "cat-punto": [
        {"codigo": "5", "descripcion": "Call Center"},
        {"codigo": "1", "descripcion": "Presencial"},
        {"codigo": "2", "descripcion": "Virtual"},
        {"codigo": "3", "descripcion": "Escrito"}
    ],
    "cat-admision": [
        {"codigo": "9", "descripcion": "No aplica"},
        {"codigo": "1", "descripcion": "Admitida"},
        {"codigo": "2", "descripcion": "No admitida"}
    ],
    "cat-ente": [
        {"codigo": "99", "descripcion": "Otros"},
        {"codigo": "1", "descripcion": "SFC"},
        {"codigo": "2", "descripcion": "Defensor CF"}
    ],
    "cat-sexo": [
        {"codigo": "M", "descripcion": "Masculino"},
        {"codigo": "F", "descripcion": "Femenino"},
        {"codigo": "I", "descripcion": "No informa"}
    ],
    "cat-lgbtiq": [
        {"codigo": "SI", "descripcion": "Pertenece a la comunidad LGBTIQ+"},
        {"codigo": "NO", "descripcion": "No pertenece"},
        {"codigo": "I", "descripcion": "No informa"}
    ],
    "cat-cond-esp": [
        {"codigo": "NINGUNA", "descripcion": "Ninguna"},
        {"codigo": "ADULTO_MAYOR", "descripcion": "Adulto mayor"},
        {"codigo": "DISCAPACIDAD_FISICA", "descripcion": "Discapacidad física"},
        {"codigo": "COGNITIVA", "descripcion": "Cognitiva"},
        {"codigo": "VULNERABLE", "descripcion": "Vulnerable"}
    ],
    "cat-prod-digital": [
        {"codigo": "1", "descripcion": "Sí"},
        {"codigo": "2", "descripcion": "No"}
    ],
    "cat-estado-queja": [
        {"codigo": "CERRADA_FAVOR_CF", "descripcion": "Cerrada a favor CF"},
        {"codigo": "CERRADA_FAVOR_ENTIDAD", "descripcion": "Cerrada a favor entidad"},
        {"codigo": "DESISTIDA", "descripcion": "Desistida"},
        {"codigo": "RECTIFICADA", "descripcion": "Rectificada"}
    ],
    "cat-favorab": [
        {"codigo": "1", "descripcion": "CF"},
        {"codigo": "2", "descripcion": "Entidad"},
        {"codigo": "3", "descripcion": "Parcial"}
    ],
    "cat-aceptacion": [
        {"codigo": "ACEPTADA_TOTAL", "descripcion": "Aceptada Total"},
        {"codigo": "ACEPTADA_PARCIAL", "descripcion": "Aceptada Parcial"},
        {"codigo": "RECHAZADA", "descripcion": "Rechazada"}
    ],
    "cat-rectif": [
        {"codigo": "RECTIFICADA", "descripcion": "Rectificada"},
        {"codigo": "NO_RECTIFICADA", "descripcion": "No Rectificada"}
    ],
    "cat-desist": [
        {"codigo": "DESISTIMIENTO_EXPRESO", "descripcion": "Desistimiento Expreso"},
        {"codigo": "DESISTIMIENTO_TACITO", "descripcion": "Desistimiento Tácito"}
    ],
    "cat-tutela": [
        {"codigo": "2", "descripcion": "No"},
        {"codigo": "1", "descripcion": "Sí"}
    ],
    "cat-marcacion": [
        {"codigo": "MARCADA", "descripcion": "Marcación especial"},
        {"codigo": "SIN_MARCACION", "descripcion": "Sin marcación"}
    ],
    "cat-expres": [
        {"codigo": "2", "descripcion": "No"},
        {"codigo": "1", "descripcion": "Sí"}
    ],
    "cat-tipo-fraude": [
        {"codigo": "FRAUDE_EXTERNO", "descripcion": "Fraude externo"},
        {"codigo": "FRAUDE_INTERNO", "descripcion": "Fraude interno"},
        {"codigo": "PHISHING", "descripcion": "Phishing"},
        {"codigo": "SUPLANTACION", "descripcion": "Suplantación"}
    ],
    "cat-mod-fraude": [
        {"codigo": "ROBO_INFO", "descripcion": "Robo de información"},
        {"codigo": "FALSIFICACION_DOCS", "descripcion": "Falsificación de documentos"},
        {"codigo": "SUPLANTACION_IDENTIDAD", "descripcion": "Suplantación de identidad"}
    ],
    "cat-area": [
        {"codigo_area": "SIN_AUTO", "nombre_area": "Siniestros Auto"},
        {"codigo_area": "SIN_VIDA", "nombre_area": "Siniestros Vida"},
        {"codigo_area": "PAGOS", "nombre_area": "Pagos"},
        {"codigo_area": "PRODUCTO", "nombre_area": "Producto"},
        {"codigo_area": "SAC", "nombre_area": "SAC"}
    ],
    "cat-usuarios-role": [
        {"codigo_area": "SIN_AUTO", "usuario": "juan.perez", "nombre_usuario": "Juan Pérez (Analista)", "rol": "Analista"},
        {"codigo_area": "SIN_AUTO", "usuario": "jorge.diaz", "nombre_usuario": "Jorge Díaz (Coordinador)", "rol": "Coordinador"},
        {"codigo_area": "SIN_VIDA", "usuario": "maria.gomez", "nombre_usuario": "María Gómez (Analista)", "rol": "Analista"},
        {"codigo_area": "PAGOS", "usuario": "carlos.rodriguez", "nombre_usuario": "Carlos Rodríguez (Coordinador)", "rol": "Coordinador"},
        {"codigo_area": "SAC", "usuario": "ana.martinez", "nombre_usuario": "Ana Martínez (Director)", "rol": "Director"},
        {"codigo_area": "PRODUCTO", "usuario": "luis.sanchez", "nombre_usuario": "Luis Sánchez (Coordinador)", "rol": "Coordinador"}
    ],
    "cat-motivo-reasig": [
        {"codigo": "ERROR_ASIGNACION", "descripcion": "Error asignación inicial"},
        {"codigo": "AREA_EQUIVOCADA", "descripcion": "Área equivocada"},
        {"codigo": "DERIVACION_PRODUCTO", "descripcion": "Derivación producto"}
    ],
    "cat-motivo-prorr": [
        {"codigo": "COMPLEJIDAD", "descripcion": "Complejidad del caso"},
        {"codigo": "FALTA_DOCUMENTACION", "descripcion": "Falta documentación del cliente"},
        {"codigo": "PROCESO_JUDICIAL", "descripcion": "Espera de proceso judicial"}
    ],
    "cat-rol-radicador": [
        {"codigo_instancia": "2", "codigo_rol_radicador": "CLIENTE", "nombre_rol_radicador": "Cliente"},
        {"codigo_instancia": "2", "codigo_rol_radicador": "INTERMEDIARIO", "nombre_rol_radicador": "Intermediario"},
        {"codigo_instancia": "2", "codigo_rol_radicador": "EMPLEADO", "nombre_rol_radicador": "Empleado Zurich"},
        {"codigo_instancia": "1", "codigo_rol_radicador": "DEFENSOR", "nombre_rol_radicador": "Defensor del Consumidor"}
    ],
    "cat-tipo-solic-pqrs": [
        {"codigo": "SOLICITUD", "descripcion": "Solicitud"},
        {"codigo": "FELICITACION", "descripcion": "Felicitación"},
        {"codigo": "QUEJA", "descripcion": "Queja"},
        {"codigo": "SUGERENCIA", "descripcion": "Sugerencia"},
        {"codigo": "PETICION", "descripcion": "Derecho de petición"}
    ],
    "cat-detalle-producto": [
        {"codigo_producto_sfc": "101", "codigo_detalle_producto": "10101", "nombre_detalle_producto": "Autos Livianos"},
        {"codigo_producto_sfc": "101", "codigo_detalle_producto": "10102", "nombre_detalle_producto": "Autos Pesados"},
        {"codigo_producto_sfc": "102", "codigo_detalle_producto": "10201", "nombre_detalle_producto": "Vida Individual"},
        {"codigo_producto_sfc": "102", "codigo_detalle_producto": "10202", "nombre_detalle_producto": "Vida Colectivo"},
        {"codigo_producto_sfc": "103", "codigo_detalle_producto": "10301", "nombre_detalle_producto": "Incendio Hogar"},
        {"codigo_producto_sfc": "103", "codigo_detalle_producto": "10302", "nombre_detalle_producto": "Terremoto Hogar"}
    ]
}


def decrypt_token(blob, key_raw):
    if not blob or blob.startswith("eyJ"):
        return blob

    # Node code for decryption
    strNodeCode = f"""
const crypto = require('crypto');
try {{
    const key = crypto.createHash('sha256').update({json.dumps(key_raw)}).digest();
    const buf = Buffer.from({json.dumps(blob)}.replace(/-/g, '+').replace(/_/g, '/'), 'base64');
    const iv = buf.subarray(0, 16);
    const ciphertext = buf.subarray(16);
    const decipher = crypto.createDecipheriv('aes-256-cbc', key, iv);
    const decrypted = Buffer.concat([decipher.update(ciphertext), decipher.final()]);
    const payload = JSON.parse(decrypted.toString('utf8'));
    console.log(payload.token);
}} catch (e) {{
    console.error(e.message);
    process.exit(1);
}}
"""
    # Try running local node first
    try:
        objRes = subprocess.run(["node", "-e", strNodeCode], capture_output=True, text=True, check=True)
        return objRes.stdout.strip()
    except (subprocess.SubprocessError, FileNotFoundError):
        # If local node fails or is not found, try via docker
        try:
            objRes = subprocess.run(["docker", "exec", "pm4-app-container", "node", "-e", strNodeCode], capture_output=True, text=True, check=True)
            return objRes.stdout.strip()
        except Exception as excError:
            print(f"Error al desencriptar el token de PM4: {excError}")
            print("Asegúrate de que 'node' local o el contenedor docker 'pm4-app-container' estén disponibles.")
            sys.exit(1)


def load_env(env_path):
    # Leemos el archivo .env y devolvemos un diccionario de variables
    dicEnvVars = {}
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as objFile:
            for strLine in objFile:
                strLine = strLine.strip()
                if strLine and not strLine.startswith("#") and "=" in strLine:
                    strKey, strValue = strLine.split("=", 1)
                    dicEnvVars[strKey.strip()] = strValue.strip()
    return dicEnvVars


def get_catalog_fields(in_strSlug):
    """Retorna los campos de base de datos / inputs de formulario requeridos según el slug."""
    if in_strSlug == "cat-dpto":
        return [
            {"name": "codigo_departamento", "label": "Código Departamento"},
            {"name": "nombre_departamento", "label": "Nombre Departamento"}
        ]
    elif in_strSlug == "cat-mpio":
        return [
            {"name": "codigo_departamento", "label": "Código Departamento (Padre)"},
            {"name": "codigo_municipio", "label": "Código Municipio"},
            {"name": "nombre_municipio", "label": "Nombre Municipio"}
        ]
    elif in_strSlug == "cat-area":
        return [
            {"name": "codigo_area", "label": "Código Área"},
            {"name": "nombre_area", "label": "Nombre Área"}
        ]
    elif in_strSlug == "cat-usuarios-role":
        return [
            {"name": "codigo_area", "label": "Código Área (Padre)"},
            {"name": "usuario", "label": "Usuario"},
            {"name": "nombre_usuario", "label": "Nombre de Usuario"},
            {"name": "rol", "label": "Rol"}
        ]
    elif in_strSlug == "cat-producto-sfc":
        return [
            {"name": "codigo_producto_sfc", "label": "Código Producto SFC"},
            {"name": "nombre_producto_sfc", "label": "Nombre Producto SFC"}
        ]
    elif in_strSlug == "cat-detalle-producto":
        return [
            {"name": "codigo_producto_sfc", "label": "Código Producto SFC (Padre)"},
            {"name": "codigo_detalle_producto", "label": "Código Detalle Producto"},
            {"name": "nombre_detalle_producto", "label": "Nombre Detalle Producto"}
        ]
    elif in_strSlug == "cat-rol-radicador":
        # CATALOGOS v2: se elimina la relación con instancia; ahora se deriva por reglas en el frontend.
        return [
            {"name": "codigo_rol_radicador", "label": "Código Rol Radicador"},
            {"name": "nombre_rol_radicador", "label": "Nombre Rol Radicador"}
        ]
    elif in_strSlug == "cat-tipo-id":
        # CATALOGOS v2: se agrega la relación con cat-tipo-persona (1 Natural / 2 Jurídica).
        return [
            {"name": "codigo", "label": "Código"},
            {"name": "descripcion", "label": "Descripción"},
            {"name": "codigo_tipo_persona", "label": "Código Tipo Persona (Relación)"}
        ]
    elif in_strSlug == "cat-alianza":
        return [
            {"name": "codigo", "label": "Código"},
            {"name": "alianza", "label": "Alianza"}
        ]
    else:
        # Catálogo estándar (con código y descripción)
        return [
            {"name": "codigo", "label": "Código"},
            {"name": "descripcion", "label": "Descripción"}
        ]


def clean_screen_title(in_strTitle):
    # Remueve cualquier caracter que no sea alfanumérico, espacio o guion
    return re.sub(r"[^a-zA-Z0-9\s\-]", "", in_strTitle).strip()


def generate_screen_json(title, fields, in_blnIsView=False, in_strScreenCatId="1"):
    """Genera la estructura JSON de una pantalla de ProcessMaker 4."""
    strCleanedTitle = clean_screen_title(title)
    lstItems = []

    # 1. Inputs del Formulario
    for dicField in fields:
        lstItems.append({
            "uuid": str(uuid.uuid4()),
            "label": "Line Input",
            "component": "FormInput",
            "config": {
                "icon": "far fa-square",
                "name": dicField["name"],
                "type": "text",
                "label": dicField["label"],
                "helper": None,
                "readonly": in_blnIsView,
                "dataFormat": "string",
                "validation": [] if in_blnIsView else ["required"],
                "placeholder": f"Ingrese {dicField['label'].lower()}",
                "defaultValue": { "mode": "js", "value": None },
                "conditionalHide": None,
                "customCssSelector": None
            },
            "editor-control": "FormInput",
            "editor-component": "FormInput"
        })

    # 2. Botón de guardar si no es pantalla de lectura
    if not in_blnIsView:
        lstItems.append({
            "uuid": str(uuid.uuid4()),
            "label": "Submit Button",
            "component": "FormButton",
            "config": {
                "icon": "fas fa-share-square",
                "name": None,
                "event": "submit",
                "label": "GRABAR",
                "loading": False,
                "tooltip": [],
                "variant": "primary",
                "fieldValue": None,
                "loadingLabel": "Loading...",
                "defaultSubmit": True
            },
            "editor-control": "FormSubmit",
            "editor-component": "FormButton"
        })

    return {
        "title": f"Ver - {strCleanedTitle}" if in_blnIsView else f"Crear - {strCleanedTitle}",
        "description": f"Pantalla de visualización para {title}" if in_blnIsView else f"Pantalla de creación/edición para {title}",
        "type": "FORM",
        "screen_category_id": in_strScreenCatId,
        "config": [
            {
                "name": title,
                "order": 1,
                "items": lstItems
            }
        ],
        "computed": [],
        "watchers": [],
        "custom_css": None,
        "status": "ACTIVE",
        "key": None,
        "is_template": 0,
        "is_default": 0
    }


def create_screen(base_url, token, screen_data):
    """POST /screens en PM4 API"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/screens"
    dicPayload = {
        "title": screen_data["title"],
        "type": screen_data["type"],
        "description": screen_data["description"],
        "config": screen_data["config"],
        "computed": screen_data.get("computed", []),
        "watchers": screen_data.get("watchers", []),
        "custom_css": screen_data.get("custom_css"),
        "screen_category_id": screen_data.get("screen_category_id")
    }
    objResp = requests.post(strUrl, json=dicPayload, headers=dicHeaders)
    objResp.raise_for_status()
    return objResp.json()


def get_existing_collections(base_url, token):
    """GET /collections para mapear slugs existentes a IDs"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/collections"
    objResp = requests.get(strUrl, params={"per_page": 500}, headers=dicHeaders)
    objResp.raise_for_status()
    dicData = objResp.json()
    return {dicCol["name"]: dicCol["id"] for dicCol in dicData.get("data", [])}


def create_collection(base_url, token, col_data):
    """POST /collections en PM4 API"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/collections"
    objResp = requests.post(strUrl, json=col_data, headers=dicHeaders)
    objResp.raise_for_status()
    return objResp.json()


def truncate_collection(base_url, token, collection_id):
    """DELETE /collections/{id}/truncate para borrar registros viejos"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/collections/{collection_id}/truncate"
    objResp = requests.delete(strUrl, headers=dicHeaders)
    objResp.raise_for_status()
    return objResp.status_code


def create_record(base_url, token, collection_id, record_data):
    """POST /collections/{id}/records en PM4 API"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/collections/{collection_id}/records"
    dicPayload = {
        "data": record_data
    }
    objResp = requests.post(strUrl, json=dicPayload, headers=dicHeaders)
    objResp.raise_for_status()
    return objResp.json()


def get_screen_category_id(base_url, token):
    """Obtiene el ID de categoría de pantalla buscando 'Fast Flow - Zurich'"""
    dicHeaders = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    strUrl = f"{base_url}/api/1.0/screen_categories"
    try:
        objResp = requests.get(strUrl, headers=dicHeaders)
        objResp.raise_for_status()
        lstCats = objResp.json().get("data", [])
        # Buscamos la categoria de Zurich y si no, usamos la primera
        for dicCat in lstCats:
            if dicCat.get("name") == "Fast Flow - Zurich":
                return str(dicCat.get("id"))
        if lstCats:
            return str(lstCats[0].get("id"))
    except Exception as excError:
        print(f"Advertencia: No se pudo obtener las categorías de screens ({excError}). Usando fallback ID '1'.")
    return "1"


def main():
    objParser = argparse.ArgumentParser(description="Script para importar catálogos de excel a PM4.")
    objParser.add_argument("--dry-run", action="store_true", help="Valida localmente la estructuración de catálogos y pantallas sin llamar a la API.")
    objArgs = objParser.parse_args()

    # Rutas relativas
    strHere = os.path.dirname(os.path.abspath(__file__))
    strExcelPath = os.path.join(strHere, "insumos", "Anexo02_Mockups_TOBE_QuejaDirectas_v3_1.xlsx")
    strEnvPath = os.path.join(strHere, ".env")

    print("=== PROCESO DE MIGRACIÓN DE CATÁLOGOS PM4 ===")
    print(f"Cargando Excel desde: {strExcelPath}")
    if not os.path.exists(strExcelPath):
        print(f"Error: No se encontró el archivo Excel en {strExcelPath}")
        sys.exit(1)

    # 1. Cargar variables de entorno si no es Dry Run
    strPm4Url = None
    strPm4Token = None
    strScreenCatId = "7" # Default para dry-run
    if not objArgs.dry_run:
        print(f"Cargando variables desde: {strEnvPath}")
        dicEnv = load_env(strEnvPath)
        strPm4Url = dicEnv.get("PM4_BASE_URL")
        strPm4Token = dicEnv.get("PM4_TOKEN")

        # Decrypt token if it is encrypted
        strPm4Token = decrypt_token(strPm4Token, dicEnv.get("IFRAME_ENCRYPTION_KEY"))

        # Obtener el ID de categoría de pantalla dinámicamente
        strScreenCatId = get_screen_category_id(strPm4Url, strPm4Token)
        print(f"Usando ID de categoría de Screen: {strScreenCatId}")

        if not strPm4Url or not strPm4Token:
            print("Error: PM4_BASE_URL y PM4_TOKEN deben estar definidos en el archivo .env")
            sys.exit(1)
        print(f"PM4 base: {strPm4Url}")

    # 2. Leer hoja de Catálogos
    objWb = openpyxl.load_workbook(strExcelPath, read_only=True, data_only=True)
    if "07_Catalogs" not in objWb.sheetnames:
        print("Error: La hoja '07_Catalogs' no existe en el libro de Excel.")
        sys.exit(1)

    objWs = objWb["07_Catalogs"]
    lstCatalogs = []

    # Empezamos en la fila 6
    intRow = 6
    while True:
        genCatId = objWs.cell(intRow, 2).value
        strName = objWs.cell(intRow, 3).value
        if not genCatId and not strName:
            # Fin del listado
            break

        if genCatId:
            strSlug = str(genCatId).strip().lower()
            lstCatalogs.append({
                "id": str(genCatId).strip(),
                "slug": strSlug,
                "nombre": str(strName).strip() if strName else str(genCatId).strip(),
                "description": objWs.cell(intRow, 5).value or ""
            })
        intRow += 1

    print(f"Se identificaron {len(lstCatalogs)} catálogos en el Excel.")

    # 3. Consultar colecciones existentes si no es dry run
    dicExistingCols = {}
    if not objArgs.dry_run:
        try:
            print("Obteniendo colecciones existentes de PM4…")
            dicExistingCols = get_existing_collections(strPm4Url, strPm4Token)
            print(f"Encontradas {len(dicExistingCols)} colecciones en la instancia.")
        except Exception as excError:
            print(f"Error al conectar con la API de PM4: {excError}")
            sys.exit(1)

    # 4. Iterar y procesar cada catálogo
    intSuccessCount = 0
    for intIdx, dicCat in enumerate(lstCatalogs):
        strSlug = dicCat["slug"]
        strName = dicCat["nombre"]
        strDesc = dicCat["description"]

        print(f"\n[{intIdx+1}/{len(lstCatalogs)}] Procesando {dicCat['id']} ('{strName}')…")

        # Obtener los campos
        lstFields = get_catalog_fields(strSlug)
        lstRecords = MOCK_RECORDS.get(strSlug, [])

        if not lstRecords:
            # Fallback generico por si falta mapeo
            print(f"  [!] Advertencia: Sin datos simulados definidos para {strSlug}. Se creará vacío.")

        print(f"  Campos definidos: {[dicField['name'] for dicField in lstFields]}")
        print(f"  Cantidad de registros sugeridos: {len(lstRecords)}")

        if objArgs.dry_run:
            # Generar JSONs para validar sintaxis de Dry Run
            dicCreateScreenJson = generate_screen_json(strName, lstFields, in_blnIsView=False, in_strScreenCatId=strScreenCatId)
            dicViewScreenJson = generate_screen_json(strName, lstFields, in_blnIsView=True, in_strScreenCatId=strScreenCatId)
            print(f"  [Dry Run] Generada Create Screen '{dicCreateScreenJson['title']}' con {len(dicCreateScreenJson['config'][0]['items'])} items.")
            print(f"  [Dry Run] Generada View Screen '{dicViewScreenJson['title']}' con {len(dicViewScreenJson['config'][0]['items'])} items.")
            intSuccessCount += 1
            continue

        # Proceso real llamando a API
        try:
            genColId = dicExistingCols.get(strSlug)

            if genColId:
                print(f"  La colección '{strSlug}' ya existe (ID: {genColId}). Limpiando registros antiguos…")
                truncate_collection(strPm4Url, strPm4Token, genColId)
                print("  Registros eliminados exitosamente.")
            else:
                # 1. Crear Screen de Creación/Edición
                print("  Creando pantalla de Creación/Edición…")
                dicCreateScreenData = generate_screen_json(strName, lstFields, in_blnIsView=False, in_strScreenCatId=strScreenCatId)
                dicCScreen = create_screen(strPm4Url, strPm4Token, dicCreateScreenData)
                genCreateScreenId = dicCScreen["id"]
                print(f"    Creada con ID: {genCreateScreenId}")

                # 2. Crear Screen de Visualización (Readonly)
                print("  Creando pantalla de Visualización (Readonly)…")
                dicViewScreenData = generate_screen_json(strName, lstFields, in_blnIsView=True, in_strScreenCatId=strScreenCatId)
                dicVScreen = create_screen(strPm4Url, strPm4Token, dicViewScreenData)
                genReadScreenId = dicVScreen["id"]
                print(f"    Creada con ID: {genReadScreenId}")

                # 3. Crear Colección
                print("  Creando la colección…")
                dicColPayload = {
                    "name": strSlug,
                    "custom_title": strName,
                    "description": strDesc or f"Colección para {strName}",
                    "create_screen_id": str(genCreateScreenId),
                    "read_screen_id": str(genReadScreenId),
                    "update_screen_id": str(genCreateScreenId),
                    "signal_create": False,
                    "signal_update": False,
                    "signal_delete": False
                }
                dicNewCol = create_collection(strPm4Url, strPm4Token, dicColPayload)
                genColId = dicNewCol["id"]
                print(f"    Colección creada con ID: {genColId}")

            # 4. Insertar los Registros
            if lstRecords:
                print(f"  Insertando {len(lstRecords)} registros…")
                intInserted = 0
                for dicRec in lstRecords:
                    create_record(strPm4Url, strPm4Token, genColId, dicRec)
                    intInserted += 1
                print(f"    Importados {intInserted}/{len(lstRecords)} registros con éxito.")

            intSuccessCount += 1
        except Exception as excError:
            print(f"  [X] ERROR procesando catálogo {dicCat['id']}: {excError}")
            if isinstance(excError, requests.exceptions.HTTPError):
                print(f"    Detalle HTTP: {excError.response.text}")

    print("\n==============================================")
    if objArgs.dry_run:
        print(f"DRY RUN COMPLETADO. Se validaron {intSuccessCount}/{len(lstCatalogs)} catálogos con éxito.")
    else:
        print(f"MIGRACIÓN COMPLETADA. Se procesaron {intSuccessCount}/{len(lstCatalogs)} catálogos con éxito.")
    print("==============================================")


if __name__ == "__main__":
    main()
